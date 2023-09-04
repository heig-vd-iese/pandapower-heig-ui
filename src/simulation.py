# Import package
import os
import logging
import coloredlogs
import traceback
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime, time, timedelta
import pandapower as pp
import pandapower.control as control
import pandapower.timeseries as timeseries
from pandapower.timeseries.data_sources.frame_data import DFData


log = logging.getLogger(__name__)
coloredlogs.install(level="INFO")

def load_net_from_xlsx(file_path: str) -> pp.pandapowerNet:
    """
    Create a pandaPower network object using data stored in a xlsx file. Pay attention to the following points:

        - Xlsx file template should be the one provided in the `package repository <https://github.com/heig-vd-iese/pandapower-heig-ui/tree/main/template>`_.
        - Every column marked in green is mandatory.
        - Every column marked in yellow is mandatory for single phase short-circuit simulation.
        - Every column marked in red is mandatory to map loads and generators to their timeseries power profile (timeseries powerflow simulation).


    INPUT:
        **file_path** (str): File path of the xlsx file where the power network data are stored.

    OUTPUT:
        **net** (pandapower.pandapowerNet): Created pandaPower object from xlsx file

    """
    # Columns which have to be set by default when None value is founded
    default_values: dict = {
        "in_service": True, "g_us_per_km": 0.0, "g0_us_per_km": 0.0, "r0_ohm_per_km": 0.0, "x0_ohm_per_km": 0.0,
        "c0_nf_per_km": 0.0, "parallel": 1, "df": 1.0, "p_mw": 0.0, "q_mvar": 0.0, "const_z_percent": 0.0,
        "const_i_percent": 0.0, "scaling": 1.0, "vk0_percent": 0.0, "vkr0_percent": 0.0, "mag0_percent": 0.0,
        "mag0_rx": 0.0, "si0_hv_partial": 0.0, "shift_degree": 0.0, "tap_step_percent": 1.0, "tap_phase_shifter": False,
        "tap_step_degree": 0.0, "vm_pu": 1.0, "va_degree": 0.0, "slack_weight": 1.0, "tap_pos": 0, "tap_neutral": 0,
        "tap_min": 0, "tap_max": 1, "profile_mapping": -1, "current_source": True
    }
    # Columns which as to be converted to another type
    int_column: list[str] = ["bus", "parallel", "from_bus", "to_bus", "element", "hv_bus", "lv_bus", "tap_pos",
                             "tap_neutral", "tap_min", "tap_max", "profile_mapping"]
    bool_column: list[str] = ["current_source", "in_service"]
    # Create empty network
    net: pp.pandapowerNet = pp.create_empty_network()
    # Open Excel file and iterate over every existing sheet table
    for eq_name in openpyxl.load_workbook(file_path).sheetnames:
        # Create a dataFrame form an Excel sheet table
        data_df: pd.DataFrame = pd.read_excel(file_path, sheet_name=eq_name).drop(columns="idx").dropna(how="all")
        if not data_df.empty:
            # Fill null values using default_values dictionary
            data_df.fillna(value=default_values, inplace=True)
            # Fill bus and load types
            if eq_name == "bus":
                data_df.fillna(value={"type": "b"}, inplace=True)
            elif eq_name == "load":
                data_df.fillna(value={"type": "wye"}, inplace=True)
            elif eq_name == "line_geodata":
                # Create list of coordinates from string
                data_df["coords"] = data_df.coords.apply(
                    lambda x: list(map(lambda y: [float(z) for z in y.split(",")],
                                       x.replace("[[", "").replace("]]", "").split("], ["))))
            # Change needed columns type from float to int64
            col: list[str] = list(set(data_df.columns).intersection(int_column))
            data_df[col] = data_df[col].astype('int64')
            # Change needed columns type from float to bool
            col: list[str] = list(set(data_df.columns).intersection(bool_column))
            data_df[col] = data_df[col].astype(bool)
            # Replace np.nan to None
            data_df = data_df.replace(np.nan, None)
            # Create pandapower network
            net[eq_name] = data_df
    return net


def load_power_profile_form_xlsx(file_path: str) -> dict[str, dict[str, pd.DataFrame]]:
    """
    Load power profile from a xlsx file and return a dictionary of pandas DataFrame ready to be applied to a
    pandaPower network for timeseries simulations. The function also interpolate power profiles in order they all have
    the same time steps. Finally, the function is only able to load one-day power profiles.

    INPUT:
        **file_path** (str): File path of the xlsx file where the power network data are stored.

    OUTPUT:
        **power_profile** (dict[str, dict[str, pd.DataFrame]]): Power profiles stored in dictionary of pandas DataFrame
        format

    """
    # Initialize power profiles dictionaries
    power_profile: dict[str, pd.DataFrame] = dict()
    results: dict[str, dict[str, pd.DataFrame]] = dict()
    # Initialize times which will be used to interpolate power profiles
    period = timedelta(days=1)
    end_time = time()
    start_time = time(23, 59, 59)
    # Load every Excel sheet names
    for eq_name in openpyxl.load_workbook(file_path).sheetnames:
        # Create a dataFrame form an Excel sheet table
        data_df = pd.read_excel(
            file_path, sheet_name=eq_name, header=[0, 1], index_col=0
        ).dropna(how="all", axis=1).dropna(how="all", axis=0)
        # If there is no data within the Excel sheet don't save the dataFrame
        if not data_df.empty:
            # Save DataFrame in power profile dictionary
            power_profile[eq_name] = data_df
            # Find the most suitable common period, start time and end time
            date_time = pd.Series(data_df.index).apply(lambda x: datetime.combine(datetime.today(), x))
            period = min(period, date_time.diff().min())
            start_time = min(start_time, data_df.index[0])
            end_time = max(end_time, data_df.index[-1])

    # Create common datetime index for every power profiles
    datetime_index = pd.Series(
        pd.date_range(
            start=datetime.combine(datetime.today(), start_time),
            end=datetime.combine(datetime.today(), end_time),
            freq=period
        ).to_pydatetime()).apply(lambda x: x.time())
    # Create a dataframe with common time and its corresponding timestamp number
    datetime_inter = pd.DataFrame(
        range(datetime_index.shape[0]), index=datetime_index,
        columns=["timestamp"]
    )
    for eq_name, eq_power in power_profile.items():
        # Create multiindex columns from power profile
        columns = pd.MultiIndex.from_tuples(
            eq_power.columns, names=["profile", "power"]
        )
        # Concatenate power profile with created common datetime index
        data_df = pd.concat([datetime_inter, eq_power], axis=1).sort_index()
        # Fill timestamps unmatched using closest neighbors (forward and backward in order to fill every null)
        data_df["timestamp"] = data_df["timestamp"].fillna(method='ffill').fillna(method='bfill')
        # Group timestamp to keep only the first founded non-null value and then fill null values using
        # first order interpolation
        data_df = data_df \
            .groupby("timestamp").first() \
            .interpolate().fillna(method="bfill").fillna(method="ffill") \
            # Create interpolated dataframe with good index and apply into results dictionary
        actual_profile = pd.DataFrame(
            data_df.values, index=datetime_inter.index, columns=columns
        )
        results[eq_name] = dict()
        if "P [MW]" in eq_power.columns.levels[1]:
            results[eq_name]["p_mw"] = actual_profile.xs("P [MW]", level=1, axis=1).dropna(how="all", axis=1)
        if "Q [MVAR]" in eq_power.columns.levels[1]:
            results[eq_name]["q_mvar"] = actual_profile.xs("Q [MVAR]", level=1, axis=1).dropna(how="all", axis=1)
    return results


def apply_power_profile(net: pp.pandapowerNet, equipment: str, power_profiles: dict[str, pd.DataFrame]):
    """
    Apply power profiles stored in dictionary of pandas DataFrame format to a pandaPower network in order to perform
    `pandaPower time simulation powerflow docs <https://pandapower.readthedocs.io/en/v2.0.1/powerflow.html>`_.
    The function create
    `pandaPower Controller <https://pandapower.readthedocs.io/en/v2.13.1/control/controller.html#constcontrol>`_ for each
    DataFrame found in power_profiles input. If the network already contains controllers linked to the equipment input,
    the function will first erase them. In order to apply wanted power profiles to wanted element, use profile_mapping
    column found in load and sgen equipments.


    INPUT:
        **net** (pandapower.pandapowerNet): pandaPower network object.

        **equipment** (str): Equipment name in which power profiles will be applied (load, sgen)

        **power_profile** (dict[str, pd.DataFrame]): Power profiles stored in dictionary of pandas DataFrame
        format. The keys of the dictionary correspond to the power parameter of equipments ("p_mw", "q_mvar")

    """
    # Check if pandapower network already have controllers for the given equipment and delete them
    if not net.controller.empty:
        old_controller_index = net.controller[net.controller.object.apply(lambda x: x.element) == equipment].index
        net.controller = net.controller.drop(old_controller_index).reset_index(drop=True)
    # Create a dictionary from profile_mapping column
    profile_mapping: dict = net[equipment] \
        .reset_index() \
        .groupby("profile_mapping")["index"] \
        .apply(list).to_dict()
    for variable, profile in power_profiles.items():
        if profile is not None:
            #
            if "time_index" in net.keys():
                if not (net["time_index"] == pd.DataFrame(profile.index, columns=["Time"])).all().all():
                    log.error("Simulation Profiles have not the same timestamps")
            else:
                net["time_index"] = pd.DataFrame(profile.index, columns=["Time"])
            if profile_mapping:
                mapped_profile = pd.DataFrame(index=profile.index)
                for profile_id, eq_list in profile_mapping.items():
                    if profile_id in profile.columns:
                        mapped_profile[eq_list] = list(zip(*[list(profile[profile_id].values)] * len(eq_list)))
            else:
                mapped_profile = profile[net[equipment].index.intersection(profile.columns)]
            unmapped_equipment = net[equipment].index.difference(mapped_profile.columns)
            if len(unmapped_equipment) != 0:
                unmapped_eq_names = ", ".join(net[equipment].loc[unmapped_equipment, "name"].values)
                log.warning("{} equipments have no {} profiles".format(unmapped_eq_names, variable))
            mapped_profile.reset_index(drop=True, inplace=True)

            control.ConstControl(net, element=equipment, element_index=mapped_profile.columns,
                                 variable=variable, data_source=DFData(mapped_profile),
                                 profile_name=mapped_profile.columns)


def create_output_writer(net: pp.pandapowerNet, add_results: [list[str] | str] = None):
    """
    Define which results will be stored as output. The function create a
    `pandaPower OutputWriter <https://pandapower.readthedocs.io/en/v2.13.1/timeseries/output_writer.html>`_
    with the following default outputs "res_bus.vm_pu", "res_line.loading_percent" and "res_trafo.loading_percent".
    Other needed outputs could be added using add_results parameters.

    INPUT:
        **net** (pandapower.pandapowerNet): pandaPower network object.

        **add_results** ([list[str] | str], None): Additional simulations results to be stored as output.

    """
    if isinstance(add_results, list):
        results = add_results + ["res_trafo.loading_percent"]
    elif isinstance(add_results, str):
        results = [add_results, "res_trafo.loading_percent"]
    else:
        results = ["res_trafo.loading_percent"]

    ow = timeseries.OutputWriter(net)

    for result in results:
        ow.log_variable(*result.split("."))


def run_time_simulation(
        net: pp.pandapowerNet, output_filename: str = None, folder: str = r"output"
) -> dict[str, pd.DataFrame]:
    """
    Run time simulation power flow to a pandaPower network where power profile has been applied. Then, the results
    will be transformed in a dictionary of DataFrame and saved in a xlsx file if wanted.


    INPUT:
        **net** (pandapower.pandapowerNet): pandaPower network object containing a
        `Controller <https://pandapower.readthedocs.io/en/v2.13.1/control/controller.html#constcontrol>`_ and a
        `OutputWriter <https://pandapower.readthedocs.io/en/v2.13.1/timeseries/output_writer.html>`_

    OPTIONAL:
        **output_filename** (str, None): File name of the xlsx file where the simulation results will be stored. If
        this parameter is not filled, the function will not save the results.

        **folder** (str, "output"): Folder name where results will be stored.

    """
    timeseries.run_timeseries(net, time_steps=range(net["time_index"].shape[0]), verbose=False)
    results_df: dict[str, pd.DataFrame] = dict()
    for key, result in net.output_writer.at[0, "object"].output.items():
        if key != "Parameters":
            results_df[key] = result.set_index(net["time_index"].Time)
            eq = key.replace("res_", "").split(".")[0]
            name_mapping = dict(zip(net[eq].index, net[eq].name))
            results_df[key] = results_df[key].rename(columns=name_mapping)

    if output_filename:
        output_file_path = os.path.join(folder, output_filename + ".xlsx")
        try:
            if not os.path.exists(folder):
                os.makedirs(folder)
            writer: pd.ExcelWriter = pd.ExcelWriter(output_file_path)
            for keys in results_df.keys():
                results_df[keys].to_excel(writer, sheet_name=keys, index_label="Time")
        except(Exception,):
            log.error(
                "Error in {} function:\n".format(traceback.extract_stack()[-1].name) +
                "Impossible to save simulation results in {} file\n".format(output_file_path) +
                "Check output file path or if file is already open in your computer"
            )
    return results_df